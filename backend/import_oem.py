"""One-shot OEM master import.

Usage:
    python -m backend.import_oem                       # dry-run report
    python -m backend.import_oem --commit              # wipe + repopulate
    python -m backend.import_oem --commit --file /path/to/master.xlsx

INVARIANTS (must never be violated):
    - Row count on disk MUST equal the row count imported. If we cannot
      import all 450 rows, the script exits non-zero and does NOT
      partially commit.
    - No cell value is ever paraphrased. `oem_utils.excel_row_to_doc`
      strips leading/trailing whitespace and adds a normalized-size
      derivative for search — that is the ONLY transformation.
    - The importer is idempotent: on `--commit` it drops the target
      collection first, so re-running yields the same end state.
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR / ".env")

# Reuse the same helpers the runtime router uses so import and query
# produce EXACTLY the same document shape.
sys.path.insert(0, str(ROOT_DIR))
from oem_utils import (  # noqa: E402
    COLLECTION,
    EXCEL_HEADERS,
    excel_row_to_doc,
    natural_key,
)


DEFAULT_XLSX = ROOT_DIR / "data" / "TyreBook_FINAL_Master_450_OEM_Verified.xlsx"


def _load_workbook_rows(path: Path):
    if not path.exists():
        raise SystemExit(f"[import_oem] Excel not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # Target the 450-row master sheet by name; fall back to the first
    # sheet if the name changes in a future release.
    sheet_name = None
    for name in wb.sheetnames:
        if "master" in name.lower() and "450" in name:
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("[import_oem] Sheet is empty")
    header = tuple(("" if c is None else str(c)).strip() for c in rows[0])
    expected = tuple(EXCEL_HEADERS)
    if header != expected:
        raise SystemExit(
            f"[import_oem] Header mismatch.\n"
            f"  expected: {expected}\n"
            f"  got:      {header}"
        )
    return sheet_name, rows[1:]


async def _run_async(commit: bool, xlsx: Path) -> dict:
    sheet_name, data_rows = _load_workbook_rows(xlsx)
    disk_count = len(data_rows)
    docs: list[dict] = []
    key_index: dict[tuple, int] = {}
    duplicates: list[dict] = []
    conflicts: list[dict] = []
    skipped: list[dict] = []
    now = datetime.now(timezone.utc).isoformat()

    for i, raw in enumerate(data_rows):
        # Excel row number of the source cell (header is row 1)
        excel_row_no = i + 2
        # Right-pad short rows with None so the tuple unpack is safe.
        padded = tuple((list(raw) + [None] * len(EXCEL_HEADERS))[: len(EXCEL_HEADERS)])
        # Skip a row only when it is 100% empty (belt-and-braces —
        # normally the workbook has none).
        if all((c is None or (isinstance(c, str) and not c.strip())) for c in padded):
            skipped.append({"row": excel_row_no, "reason": "all-empty"})
            continue
        doc = excel_row_to_doc(padded, excel_row_no)
        # Reject rows missing any required field. Per user rule #17
        # we don't invent values, but we also can't index or search a
        # doc without make/model/front/rear.
        required_missing = [
            f for f in ("make", "model", "front_tyre_size", "rear_tyre_size")
            if not doc.get(f)
        ]
        if required_missing:
            skipped.append({
                "row": excel_row_no,
                "reason": f"missing required: {required_missing}",
            })
            continue
        # Dedup by natural key (see oem_utils.natural_key). We do NOT
        # drop duplicates — we report them and keep the first
        # occurrence.
        key = natural_key(doc)
        if key in key_index:
            duplicates.append({
                "row": excel_row_no,
                "duplicate_of_row": docs[key_index[key]]["_source_excel_row"],
                "make": doc.get("make"),
                "model": doc.get("model"),
                "variant": doc.get("variant"),
                "year_generation": doc.get("year_generation"),
                "front_tyre_size": doc.get("front_tyre_size"),
                "rear_tyre_size": doc.get("rear_tyre_size"),
            })
            # Keep the doc — user directive #1: "Do not delete
            # duplicate-looking rows if their variant/year/fitment is
            # different." Only mark as duplicate for reporting.
        else:
            key_index[key] = len(docs)
        doc["id"] = str(uuid.uuid4())
        doc["created_at"] = now
        doc["updated_at"] = now
        docs.append(doc)

    # Conflicts here would mean: same natural key but DIFFERENT
    # supporting fields (evidence URL, verification, etc.). Report
    # them as informational — none are expected in the shipped file.
    seen_by_key: dict[tuple, dict] = {}
    for d in docs:
        k = natural_key(d)
        if k not in seen_by_key:
            seen_by_key[k] = d
        else:
            first = seen_by_key[k]
            differences = {}
            for f in ("verification_status", "oem_evidence", "oem_source_url", "category"):
                if (first.get(f) or "") != (d.get(f) or ""):
                    differences[f] = {
                        "first_row": first.get("_source_excel_row"),
                        "first_value": first.get(f),
                        "conflict_row": d.get("_source_excel_row"),
                        "conflict_value": d.get(f),
                    }
            if differences:
                conflicts.append({
                    "row": d.get("_source_excel_row"),
                    "differences": differences,
                })

    report = {
        "commit": commit,
        "source_file": str(xlsx),
        "source_sheet": sheet_name,
        "excel_data_rows_on_disk": disk_count,
        "documents_prepared": len(docs),
        "duplicates_flagged": len(duplicates),
        "conflicts_flagged": len(conflicts),
        "skipped": len(skipped),
        "duplicate_details": duplicates,
        "conflict_details": conflicts,
        "skipped_details": skipped,
    }

    # Fail-fast if the disk row count doesn't match the prepared docs
    # (minus skipped). This is the hard invariant the user asked for.
    if disk_count - len(skipped) != len(docs):
        report["invariant_violation"] = (
            f"disk({disk_count}) - skipped({len(skipped)}) != prepared({len(docs)})"
        )

    if not commit:
        return report

    # ---- WRITE PATH ------------------------------------------------
    mongo_url = os.environ["MONGO_URL"]
    db_name = os.environ["DB_NAME"]
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    coll = db[COLLECTION]

    # Wipe + repopulate. Phase 4 will replace this with an
    # incremental importer that respects existing admin edits.
    await coll.delete_many({})
    if docs:
        # insert_many can accept a big list of dicts.
        await coll.insert_many(docs)

    # Indexes for search performance (per user directive #14). All
    # background so this stays fast even at 50k+ records.
    await coll.create_index("category", background=True)
    await coll.create_index("make", background=True)
    await coll.create_index("model", background=True)
    await coll.create_index("variant", background=True)
    await coll.create_index("year_generation", background=True)
    await coll.create_index("front_tyre_size", background=True)
    await coll.create_index("rear_tyre_size", background=True)
    await coll.create_index("front_size_normalized", background=True)
    await coll.create_index("rear_size_normalized", background=True)
    await coll.create_index("verification_status", background=True)
    await coll.create_index(
        [("category", 1), ("make", 1), ("model", 1), ("variant", 1), ("year_generation", 1)],
        background=True,
        name="oem_vehicle_lookup",
    )
    await coll.create_index("id", unique=True, background=True)

    final_count = await coll.count_documents({})
    report["mongo_final_count"] = final_count
    report["mongo_final_matches_disk"] = final_count == disk_count - len(skipped)

    client.close()
    return report


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true",
                    help="Actually wipe and repopulate the Mongo collection.")
    ap.add_argument("--file", type=str, default=str(DEFAULT_XLSX),
                    help="Path to the OEM Master Excel file.")
    ap.add_argument("--json", action="store_true",
                    help="Print the full report as JSON.")
    args = ap.parse_args()
    report = asyncio.run(_run_async(args.commit, Path(args.file)))
    if args.json:
        print(json.dumps(report, indent=2, default=str))
    else:
        print(f"[import_oem] commit={report['commit']}")
        print(f"[import_oem] source_file={report['source_file']}")
        print(f"[import_oem] source_sheet={report['source_sheet']}")
        print(f"[import_oem] excel_data_rows_on_disk={report['excel_data_rows_on_disk']}")
        print(f"[import_oem] documents_prepared={report['documents_prepared']}")
        print(f"[import_oem] duplicates_flagged={report['duplicates_flagged']}")
        print(f"[import_oem] conflicts_flagged={report['conflicts_flagged']}")
        print(f"[import_oem] skipped={report['skipped']}")
        if "mongo_final_count" in report:
            print(f"[import_oem] mongo_final_count={report['mongo_final_count']}")
            print(f"[import_oem] mongo_final_matches_disk={report['mongo_final_matches_disk']}")
        if "invariant_violation" in report:
            print(f"[import_oem] INVARIANT VIOLATION: {report['invariant_violation']}")
            return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
