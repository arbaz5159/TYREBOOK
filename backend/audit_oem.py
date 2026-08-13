"""OEM data-integrity audit.

Compares the live Mongo `oem_vehicle_fitments` collection against the
original 450-row master Excel (`/app/backend/data/TyreBook_FINAL_Master_450_OEM_Verified.xlsx`)
and reports:

  * total document count on disk vs. Mongo
  * how many rows match byte-for-byte on the "identity" fields
    (make, model, variant, year, front, rear, category, verification,
    evidence, source URL)
  * any Mongo doc that has no counterpart in the Excel
  * any Excel row that has no counterpart in Mongo
  * any per-field mismatches, listing the changed keys

The audit is READ-ONLY.  It never mutates the collection.

Usage:
    python -m backend.audit_oem
    python -m backend.audit_oem --json > /tmp/oem_audit.json
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
from collections import Counter
from pathlib import Path

from dotenv import load_dotenv
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

ROOT_DIR = Path(__file__).resolve().parent
load_dotenv(ROOT_DIR / ".env")
sys.path.insert(0, str(ROOT_DIR))
from oem_utils import COLLECTION, EXCEL_HEADERS, excel_row_to_doc, natural_key  # noqa: E402

DEFAULT_XLSX = ROOT_DIR / "data" / "TyreBook_FINAL_Master_450_OEM_Verified.xlsx"

IDENTITY_FIELDS = (
    "make",
    "model",
    "variant",
    "year_generation",
    "front_tyre_size",
    "rear_tyre_size",
    "category",
    "verification_status",
    "oem_evidence",
    "oem_source_url",
)


def _load_excel(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet = None
    for name in wb.sheetnames:
        if "master" in name.lower() and "450" in name:
            sheet = name
            break
    if sheet is None:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = tuple(("" if c is None else str(c)).strip() for c in rows[0])
    assert header == tuple(EXCEL_HEADERS), f"header drift: got {header}"
    out: list[dict] = []
    for i, r in enumerate(rows[1:]):
        padded = tuple((list(r) + [None] * len(EXCEL_HEADERS))[: len(EXCEL_HEADERS)])
        if all((c is None or (isinstance(c, str) and not c.strip())) for c in padded):
            continue
        doc = excel_row_to_doc(padded, i + 2)
        out.append(doc)
    return out


async def _run_async(as_json: bool) -> int:
    excel = _load_excel(DEFAULT_XLSX)
    disk_by_key: dict[tuple, dict] = {natural_key(d): d for d in excel}

    mongo_url = os.environ["MONGO_URL"]
    db_name = os.environ["DB_NAME"]
    client = AsyncIOMotorClient(mongo_url)
    db = client[db_name]
    coll = db[COLLECTION]

    mongo_docs: list[dict] = []
    async for d in coll.find({}, {"_id": 0}):
        mongo_docs.append(d)
    client.close()
    mongo_by_key: dict[tuple, dict] = {natural_key(d): d for d in mongo_docs}

    # -----------------------------------------------
    identical = 0
    field_mismatches: list[dict] = []
    only_in_mongo: list[dict] = []
    only_in_excel: list[dict] = []
    key_conflicts: list[dict] = []  # multiple docs with same natural_key

    # duplicate natural keys within Mongo
    #
    # KNOWN LEGITIMATE DUPLICATES (documented in iteration_24 test
    # report): the three TVS iQube rows share a natural_key because the
    # Variant column is blank in the source Excel, but each row points
    # to a different battery-variant OEM URL (2.2 / 3.1 / 3.5 kWh).
    # Directive #1 forbids deleting them, so their appearance in
    # key_conflicts is informational — not a data-integrity failure.
    KNOWN_LEGIT_KEYS = {
        (
            "tvs",
            "iqube",
            "",       # variant blank
            "current india model",
            "90/90-12",
            "90/90-12",
        ),
    }
    key_counts_mongo: Counter[tuple] = Counter(natural_key(d) for d in mongo_docs)
    unexpected_key_conflicts: list[dict] = []
    for k, c in key_counts_mongo.items():
        if c > 1:
            entry = {"natural_key": list(k), "count": c, "source": "mongo"}
            if k in KNOWN_LEGIT_KEYS:
                entry["known_legitimate"] = True
                key_conflicts.append(entry)
            else:
                entry["known_legitimate"] = False
                key_conflicts.append(entry)
                unexpected_key_conflicts.append(entry)

    for key, disk_doc in disk_by_key.items():
        mongo_doc = mongo_by_key.get(key)
        if mongo_doc is None:
            only_in_excel.append({"natural_key": list(key)})
            continue
        diffs = []
        for f in IDENTITY_FIELDS:
            a = disk_doc.get(f)
            b = mongo_doc.get(f)
            # normalise blank/None comparison because the importer
            # emits None for blank cells but a hand-edit could have
            # left "".
            if (a in (None, "")) and (b in (None, "")):
                continue
            if a != b:
                diffs.append({"field": f, "excel": a, "mongo": b})
        if not diffs:
            identical += 1
        else:
            field_mismatches.append({
                "natural_key": list(key),
                "diffs": diffs,
            })

    for key in mongo_by_key.keys() - disk_by_key.keys():
        only_in_mongo.append({"natural_key": list(key)})

    passed = (
        len(mongo_docs) == 450
        and len(excel) == 450
        and not field_mismatches
        and not only_in_mongo
        and not only_in_excel
        and not unexpected_key_conflicts
    )

    report = {
        "pass": passed,
        "excel_source": str(DEFAULT_XLSX),
        "excel_data_rows": len(excel),
        "mongo_documents": len(mongo_docs),
        "identical_rows": identical,
        "field_mismatches": field_mismatches,
        "only_in_mongo": only_in_mongo,
        "only_in_excel": only_in_excel,
        "key_conflicts": key_conflicts,
    }

    if as_json:
        print(json.dumps(report, indent=2, default=str))
    else:
        print(f"[audit] excel_data_rows       = {report['excel_data_rows']}")
        print(f"[audit] mongo_documents       = {report['mongo_documents']}")
        print(f"[audit] identical_rows        = {report['identical_rows']}")
        print(f"[audit] field_mismatches      = {len(field_mismatches)}")
        print(f"[audit] only_in_mongo         = {len(only_in_mongo)}")
        print(f"[audit] only_in_excel         = {len(only_in_excel)}")
        print(f"[audit] key_conflicts         = {len(key_conflicts)}")
        print(f"[audit] PASS                  = {passed}")
    return 0 if passed else 2


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", action="store_true")
    args = ap.parse_args()
    return asyncio.run(_run_async(args.json))


if __name__ == "__main__":
    raise SystemExit(main())
