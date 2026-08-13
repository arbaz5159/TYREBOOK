"""OEM tyre fitment read-only query endpoints.

Phase 1 delivers ONLY read/search endpoints — they are safe to expose
to every authenticated app user (Shop Admin + Staff). Admin write
endpoints (edit, Excel-import) will be added in Phase 4 with Firebase
ID-token verification against the Super Admin allow-list.

All endpoints:
    GET /api/oem/health
    GET /api/oem/stats
    GET /api/oem/categories
    GET /api/oem/makes?category=...
    GET /api/oem/models?category=...&make=...
    GET /api/oem/variants?category=...&make=...&model=...
    GET /api/oem/years?category=...&make=...&model=...&variant=...
    GET /api/oem/fitments?category=...&make=...&model=...&variant=...&year_generation=...
    GET /api/oem/search-by-size?size=...
    GET /api/oem/admin/list?...   (Phase 1 exposes read-only; write is Phase 4)

Design notes:
    * We deliberately DO NOT paginate distinct-value endpoints — the
      OEM master has ~450 rows today and even at 50k the distinct
      Make list is still under a few hundred. The admin list is
      paginated because it returns every row.
    * "Uncategorised" (per user directive #3) is a client-visible label
      for the subset whose source `category` cell is None. Callers pass
      the literal string "Uncategorised" and we translate it to a
      `{"$or": [{"category": None}, {"category": ""}]}` filter.
    * Size normalisation is done in `oem_utils.normalize_size`. The
      exact original size string is always returned to the UI —
      normalisation is used ONLY for the query.
"""

from __future__ import annotations

import os
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Query, Request
from motor.motor_asyncio import AsyncIOMotorDatabase

from oem_auth import require_super_admin
from oem_utils import (
    BLANK_SENTINEL,
    COLLECTION,
    EXCEL_HEADERS,
    UNCATEGORISED_LABEL,
    excel_row_to_doc,
    natural_key,
    normalize_size,
    public_projection,
)

AUDIT_COLLECTION = "oem_audit_log"

# Fields that a Super Admin is allowed to change via the edit endpoint.
# Everything else (audit metadata, derived normalisation, uuid) is
# managed by the backend.
EDITABLE_FIELDS = {
    "category",
    "make",
    "model",
    "variant",
    "year_generation",
    "front_tyre_size",
    "rear_tyre_size",
    "verification_status",
    "oem_evidence",
    "oem_source_url",
    "source_pass",
    "no",
}


def build_router(db: AsyncIOMotorDatabase) -> APIRouter:
    router = APIRouter(prefix="/oem", tags=["oem"])
    coll = db[COLLECTION]

    # ------------------------------------------------------------
    # Filter helpers
    # ------------------------------------------------------------
    def _category_filter(category: Optional[str]) -> Dict[str, Any]:
        """`Uncategorised` maps to blank/None category rows. Any other
        value is an exact match. `None`/`""` means "no filter"."""
        if category is None or category == "":
            return {}
        if category == UNCATEGORISED_LABEL:
            return {"$or": [{"category": None}, {"category": ""}]}
        return {"category": category}

    def _exact(field: str, value: Optional[str]) -> Dict[str, Any]:
        if value is None or value == "":
            return {}
        # Sentinel — the client wants rows whose source cell is blank.
        # Match both `None` and `""` because the importer preserves the
        # cell verbatim (post-`.strip()`) but Excel may deliver either.
        if value == BLANK_SENTINEL:
            return {"$or": [{field: None}, {field: ""}]}
        return {field: value}

    def _combine(*parts: Dict[str, Any]) -> Dict[str, Any]:
        merged: Dict[str, Any] = {}
        for p in parts:
            if not p:
                continue
            # If merging two $or clauses we need $and to combine safely.
            if "$or" in merged and "$or" in p:
                merged.setdefault("$and", []).append({"$or": merged.pop("$or")})
                merged["$and"].append({"$or": p["$or"]})
            elif "$or" in merged and "$or" not in p:
                merged.setdefault("$and", []).append({"$or": merged.pop("$or")})
                merged["$and"].append(p)
            elif "$or" in p and "$or" not in merged:
                if merged:
                    merged = {"$and": [merged, {"$or": p["$or"]}]}
                else:
                    merged["$or"] = p["$or"]
            else:
                merged.update(p)
        return merged

    async def _distinct(field: str, filt: Dict[str, Any]) -> List[str]:
        raw = await coll.distinct(field, filt)
        # Normalise None → UNCATEGORISED_LABEL for the "category" field
        # only. For every other field, preserve blank cells by emitting
        # the BLANK_SENTINEL so the row remains reachable from the
        # picker — the UI displays it as "(not specified)".
        if field == "category":
            out: List[str] = []
            for v in raw:
                if v is None or v == "":
                    out.append(UNCATEGORISED_LABEL)
                else:
                    out.append(v)
        else:
            out = []
            has_blank = False
            for v in raw:
                if v is None or v == "":
                    has_blank = True
                else:
                    out.append(v)
            if has_blank:
                out.append(BLANK_SENTINEL)
        # Deduplicate + case-insensitive sort. The BLANK_SENTINEL sorts
        # to the top because it starts with '_'.
        seen: Dict[str, str] = {}
        for v in out:
            key = v.lower()
            if key not in seen:
                seen[key] = v
        return sorted(seen.values(), key=lambda x: x.lower())

    # ------------------------------------------------------------
    # /health, /stats
    # ------------------------------------------------------------
    @router.get("/health")
    async def health():
        try:
            total = await coll.count_documents({})
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"db error: {e}")
        return {
            "ok": True,
            "collection": COLLECTION,
            "total_documents": total,
        }

    @router.get("/stats")
    async def stats():
        total = await coll.count_documents({})
        oem_verified = await coll.count_documents({"verification_status": "OEM VERIFIED"})
        uncategorised = await coll.count_documents(
            {"$or": [{"category": None}, {"category": ""}]}
        )
        # per-category counts
        pipeline = [
            {"$group": {"_id": "$category", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
        ]
        buckets = []
        async for row in coll.aggregate(pipeline):
            label = row["_id"] if row["_id"] not in (None, "") else UNCATEGORISED_LABEL
            buckets.append({"category": label, "count": row["count"]})
        return {
            "total": total,
            "oem_verified": oem_verified,
            "uncategorised": uncategorised,
            "by_category": buckets,
        }

    # ------------------------------------------------------------
    # Dropdown cascade
    # ------------------------------------------------------------
    @router.get("/categories")
    async def list_categories():
        values = await _distinct("category", {})
        return {"categories": values}

    @router.get("/makes")
    async def list_makes(category: Optional[str] = Query(default=None)):
        filt = _category_filter(category)
        values = await _distinct("make", filt)
        return {"makes": values}

    @router.get("/models")
    async def list_models(
        category: Optional[str] = Query(default=None),
        make: Optional[str] = Query(default=None),
    ):
        filt = _combine(_category_filter(category), _exact("make", make))
        values = await _distinct("model", filt)
        return {"models": values}

    @router.get("/variants")
    async def list_variants(
        category: Optional[str] = Query(default=None),
        make: Optional[str] = Query(default=None),
        model: Optional[str] = Query(default=None),
    ):
        filt = _combine(
            _category_filter(category),
            _exact("make", make),
            _exact("model", model),
        )
        values = await _distinct("variant", filt)
        return {"variants": values}

    @router.get("/years")
    async def list_years(
        category: Optional[str] = Query(default=None),
        make: Optional[str] = Query(default=None),
        model: Optional[str] = Query(default=None),
        variant: Optional[str] = Query(default=None),
    ):
        filt = _combine(
            _category_filter(category),
            _exact("make", make),
            _exact("model", model),
            _exact("variant", variant),
        )
        values = await _distinct("year_generation", filt)
        return {"years": values}

    # ------------------------------------------------------------
    # Fitment resolution
    # ------------------------------------------------------------
    @router.get("/fitments")
    async def list_fitments(
        category: Optional[str] = Query(default=None),
        make: Optional[str] = Query(default=None),
        model: Optional[str] = Query(default=None),
        variant: Optional[str] = Query(default=None),
        year_generation: Optional[str] = Query(default=None),
    ):
        """Return every OEM fitment matching the filters. The Vehicle
        Search UI is expected to keep prompting until a unique row
        remains (per user directives #4 and #11). If more than one row
        comes back the UI must NOT auto-select — it must ask the user
        to disambiguate.
        """
        filt = _combine(
            _category_filter(category),
            _exact("make", make),
            _exact("model", model),
            _exact("variant", variant),
            _exact("year_generation", year_generation),
        )
        rows: List[Dict[str, Any]] = []
        cursor = coll.find(filt, public_projection()).sort(
            [("make", 1), ("model", 1), ("variant", 1), ("year_generation", 1)]
        )
        async for r in cursor:
            rows.append(r)
        return {
            "count": len(rows),
            "is_unique": len(rows) == 1,
            "fitments": rows,
        }

    # ------------------------------------------------------------
    # Search by tyre size (both front and rear)
    # ------------------------------------------------------------
    @router.get("/search-by-size")
    async def search_by_size(
        size: str = Query(..., min_length=2, description="Tyre size in any common format"),
    ):
        norm = normalize_size(size)
        if not norm:
            raise HTTPException(status_code=400, detail="empty size")
        # `$or` because the same size can appear as either front or rear.
        filt = {
            "$or": [
                {"front_size_normalized": norm},
                {"rear_size_normalized": norm},
            ]
        }
        rows = []
        cursor = coll.find(filt, public_projection()).sort([("make", 1), ("model", 1)])
        async for r in cursor:
            rows.append(r)
        return {
            "query": size,
            "normalized": norm,
            "count": len(rows),
            "vehicles": rows,
        }

    # ------------------------------------------------------------
    # Admin (read-only in Phase 1 — Phase 4 will add write)
    # ------------------------------------------------------------
    @router.get("/admin/list")
    async def admin_list(
        category: Optional[str] = Query(default=None),
        make: Optional[str] = Query(default=None),
        model: Optional[str] = Query(default=None),
        size: Optional[str] = Query(default=None),
        needs_review: Optional[bool] = Query(default=None,
            description="If true, return only rows whose Category is blank."),
        page: int = Query(default=1, ge=1),
        page_size: int = Query(default=50, ge=1, le=500),
    ):
        parts: List[Dict[str, Any]] = []
        if category:
            parts.append(_category_filter(category))
        if make:
            parts.append({"make": make})
        if model:
            parts.append({"model": model})
        if size:
            norm = normalize_size(size)
            parts.append({
                "$or": [
                    {"front_size_normalized": norm},
                    {"rear_size_normalized": norm},
                ]
            })
        if needs_review is True:
            parts.append({"_category_review_required": True})
        elif needs_review is False:
            parts.append({"_category_review_required": {"$ne": True}})
        filt = _combine(*parts)

        total = await coll.count_documents(filt)
        skip = (page - 1) * page_size
        # Admin listing exposes _category_review_required so the UI can
        # badge rows, so we do NOT hide it via the projection.
        proj = {"_id": 0, "_source_excel_row": 0}
        rows = []
        cursor = coll.find(filt, proj).sort(
            [("make", 1), ("model", 1), ("variant", 1), ("year_generation", 1)]
        ).skip(skip).limit(page_size)
        async for r in cursor:
            rows.append(r)
        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "rows": rows,
        }

    @router.get("/admin/{fitment_id}")
    async def admin_get(fitment_id: str):
        row = await coll.find_one({"id": fitment_id}, {"_id": 0})
        if not row:
            raise HTTPException(status_code=404, detail="OEM fitment not found")
        return row

    # ================================================================
    # SUPER-ADMIN WRITE ENDPOINTS
    # ================================================================
    # Every write below is gated by `require_super_admin` (Firebase ID
    # token verified via public JWKS + email allow-list). All writes
    # append an entry to `oem_audit_log` — see AUDIT_COLLECTION.

    audit_coll = db[AUDIT_COLLECTION]

    async def _write_audit(
        actor: Dict[str, Any],
        action: str,
        fitment_id: Optional[str],
        before: Optional[Dict[str, Any]],
        after: Optional[Dict[str, Any]],
        changed_fields: Optional[List[str]] = None,
        extra: Optional[Dict[str, Any]] = None,
    ):
        await audit_coll.insert_one({
            "id": str(uuid.uuid4()),
            "action": action,
            "fitment_id": fitment_id,
            "actor_uid": actor.get("uid"),
            "actor_email": actor.get("email"),
            "changed_fields": changed_fields or [],
            "before": before,
            "after": after,
            "extra": extra or {},
            "timestamp": datetime.now(timezone.utc).isoformat(),
        })

    @router.put("/admin/{fitment_id}")
    async def admin_update(fitment_id: str, request: Request):
        # Verify caller BEFORE parsing the body — never leak schema
        # hints to unauthenticated callers.
        actor = require_super_admin(request)
        try:
            payload: Dict[str, Any] = await request.json()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid JSON body")
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Body must be a JSON object")

        # Prevent silent overwrites of unrelated fields — only known
        # editable fields make it into the update patch.
        patch: Dict[str, Any] = {}
        for k, v in payload.items():
            if k in EDITABLE_FIELDS:
                # Normalise `category`: blank / whitespace → None so
                # the "Uncategorised" bucket stays consistent.
                if k == "category":
                    s = (v or "").strip() if isinstance(v, str) else None
                    patch[k] = s if s else None
                elif k in ("no", "source_pass"):
                    try:
                        patch[k] = int(v) if v not in ("", None) else None
                    except (TypeError, ValueError):
                        raise HTTPException(status_code=400, detail=f"{k} must be an integer")
                else:
                    patch[k] = (v or "").strip() if isinstance(v, str) else v

        if not patch:
            raise HTTPException(status_code=400, detail="No editable fields present in body")

        existing = await coll.find_one({"id": fitment_id}, {"_id": 0})
        if not existing:
            raise HTTPException(status_code=404, detail="OEM fitment not found")

        # Compute only actual diffs so no-op writes don't spam the audit log.
        changed = [
            k for k, v in patch.items() if existing.get(k) != v
        ]
        if not changed:
            return {"ok": True, "unchanged": True, "fitment": existing}

        # Refresh derived normalisation whenever a tyre size changed.
        derived: Dict[str, Any] = {}
        if "front_tyre_size" in patch:
            derived["front_size_normalized"] = normalize_size(patch["front_tyre_size"])
        if "rear_tyre_size" in patch:
            derived["rear_size_normalized"] = normalize_size(patch["rear_tyre_size"])
        derived["_category_review_required"] = (
            patch["category"] is None if "category" in patch else existing.get("_category_review_required", False)
        )
        derived["updated_at"] = datetime.now(timezone.utc).isoformat()

        await coll.update_one({"id": fitment_id}, {"$set": {**patch, **derived}})
        after = await coll.find_one({"id": fitment_id}, {"_id": 0})

        await _write_audit(
            actor,
            action="edit",
            fitment_id=fitment_id,
            before={k: existing.get(k) for k in changed},
            after={k: (after or {}).get(k) for k in changed},
            changed_fields=changed,
        )
        return {"ok": True, "changed": changed, "fitment": after}

    # ---------------- Excel import (Super Admin) ------------------
    @router.post("/admin/import/preview")
    async def admin_import_preview(request: Request):
        actor = require_super_admin(request)
        form = await request.form()
        upload = form.get("file")
        if upload is None or not hasattr(upload, "read"):
            raise HTTPException(status_code=400, detail="Multipart 'file' field required")

        content = await upload.read()  # type: ignore[attr-defined]
        report = _analyze_workbook(content, existing_by_key=None)
        # Read existing for conflict / new detection
        existing_by_key: Dict[tuple, Dict[str, Any]] = {}
        async for doc in coll.find({}, {"_id": 0}):
            existing_by_key[natural_key(doc)] = doc
        report = _analyze_workbook(content, existing_by_key=existing_by_key)
        # Note: `actor` intentionally not logged for a preview.
        report["actor_email"] = actor.get("email")
        return report

    @router.post("/admin/import/commit")
    async def admin_import_commit(
        request: Request,
        overwrite_conflicts: bool = Query(default=False),
    ):
        actor = require_super_admin(request)
        form = await request.form()
        upload = form.get("file")
        if upload is None or not hasattr(upload, "read"):
            raise HTTPException(status_code=400, detail="Multipart 'file' field required")
        content = await upload.read()  # type: ignore[attr-defined]

        existing_by_key: Dict[tuple, Dict[str, Any]] = {}
        async for doc in coll.find({}, {"_id": 0}):
            existing_by_key[natural_key(doc)] = doc

        report = _analyze_workbook(content, existing_by_key=existing_by_key)
        if report.get("invalid_rows"):
            raise HTTPException(status_code=400, detail={
                "message": "File has invalid rows; fix them and re-upload.",
                "invalid_rows": report["invalid_rows"][:20],
            })

        # Per user directive #4 & #7: never overwrite existing records
        # unless the Super Admin explicitly ticked the confirmation.
        added = 0
        overwritten = 0
        skipped_conflicts = 0
        now = datetime.now(timezone.utc).isoformat()

        for prep in report["prepared_rows"]:
            key = tuple(prep["natural_key"])
            doc = prep["doc"]
            if key in existing_by_key:
                if overwrite_conflicts:
                    existing = existing_by_key[key]
                    doc["id"] = existing["id"]           # preserve id
                    doc["created_at"] = existing.get("created_at", now)
                    doc["updated_at"] = now
                    await coll.replace_one({"id": existing["id"]}, doc)
                    await _write_audit(
                        actor,
                        action="import-overwrite",
                        fitment_id=existing["id"],
                        before=existing,
                        after=doc,
                        changed_fields=[
                            k for k in doc.keys()
                            if existing.get(k) != doc.get(k)
                            and k not in ("_source_excel_row",)
                        ],
                    )
                    overwritten += 1
                else:
                    skipped_conflicts += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["created_at"] = now
                doc["updated_at"] = now
                await coll.insert_one(doc)
                await _write_audit(
                    actor,
                    action="import-insert",
                    fitment_id=doc["id"],
                    before=None,
                    after=doc,
                    changed_fields=list(doc.keys()),
                )
                added += 1

        final_count = await coll.count_documents({})
        return {
            "ok": True,
            "added": added,
            "overwritten": overwritten,
            "skipped_conflicts_awaiting_review": skipped_conflicts,
            "invalid_rows": report.get("invalid_rows", []),
            "final_count": final_count,
        }

    @router.get("/admin/audit-log/list")
    async def audit_log_list(
        request: Request,
        fitment_id: Optional[str] = Query(default=None),
        page: int = Query(default=1, ge=1),
        page_size: int = Query(default=50, ge=1, le=500),
    ):
        # Read requires super-admin — audit contents include email PII.
        require_super_admin(request)
        q: Dict[str, Any] = {}
        if fitment_id:
            q["fitment_id"] = fitment_id
        total = await audit_coll.count_documents(q)
        cursor = audit_coll.find(q, {"_id": 0}).sort("timestamp", -1).skip((page - 1) * page_size).limit(page_size)
        rows = []
        async for r in cursor:
            rows.append(r)
        return {"total": total, "page": page, "page_size": page_size, "rows": rows}

    return router


# ------------------------------------------------------------------
# Excel-workbook analysis helper (used by preview + commit).
# ------------------------------------------------------------------
def _analyze_workbook(
    content: bytes,
    existing_by_key: Optional[Dict[tuple, Dict[str, Any]]],
) -> Dict[str, Any]:
    """Read the uploaded .xlsx bytes and produce a validation report.

    Behaviour matches the strict rules from `import_oem.py`:
      * Header row MUST match EXCEL_HEADERS exactly.
      * Rows missing any of make / model / front_tyre_size /
        rear_tyre_size are `invalid_rows` — they never enter the
        prepared list.
      * A row whose natural_key is already in the collection is
        classified as a `conflict` (differences noted) or `duplicate`
        (byte-identical row).
      * A row whose natural_key isn't in the collection is `new`.
    """
    import io as _io  # local import so the module load is cheap.
    import openpyxl as _openpyxl

    try:
        wb = _openpyxl.load_workbook(_io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:  # noqa: BLE001
        return {
            "ok": False,
            "error": f"Cannot open Excel file: {e}",
            "prepared_rows": [],
            "invalid_rows": [],
            "conflicts": [],
            "duplicates": [],
            "new_rows": [],
        }

    sheet_name = None
    for name in wb.sheetnames:
        if "master" in name.lower():
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = tuple(("" if c is None else str(c)).strip() for c in (rows[0] if rows else ()))
    if header != tuple(EXCEL_HEADERS):
        return {
            "ok": False,
            "error": "Header mismatch — file must contain the same 12 columns as the master.",
            "expected_headers": list(EXCEL_HEADERS),
            "got_headers": list(header),
            "prepared_rows": [],
            "invalid_rows": [],
            "conflicts": [],
            "duplicates": [],
            "new_rows": [],
        }

    prepared: List[Dict[str, Any]] = []
    invalid: List[Dict[str, Any]] = []
    conflicts: List[Dict[str, Any]] = []
    duplicates: List[Dict[str, Any]] = []
    news: List[Dict[str, Any]] = []
    seen_incoming: Dict[tuple, int] = {}

    for i, raw in enumerate(rows[1:]):
        excel_row_no = i + 2
        padded = tuple((list(raw) + [None] * len(EXCEL_HEADERS))[: len(EXCEL_HEADERS)])
        if all(
            (c is None or (isinstance(c, str) and not c.strip())) for c in padded
        ):
            continue
        doc = excel_row_to_doc(padded, excel_row_no)
        missing = [
            f
            for f in ("make", "model", "front_tyre_size", "rear_tyre_size")
            if not doc.get(f)
        ]
        if missing:
            invalid.append({"row": excel_row_no, "missing": missing})
            continue
        key = natural_key(doc)
        # Dedup within the incoming file.
        if key in seen_incoming:
            duplicates.append({
                "row": excel_row_no,
                "duplicate_of_row": seen_incoming[key],
                "make": doc.get("make"),
                "model": doc.get("model"),
                "variant": doc.get("variant"),
            })
        else:
            seen_incoming[key] = excel_row_no

        if existing_by_key is not None and key in existing_by_key:
            existing = existing_by_key[key]
            diff_fields = [
                f
                for f in (
                    "category", "verification_status", "oem_evidence",
                    "oem_source_url", "source_pass", "no",
                )
                if (existing.get(f) or None) != (doc.get(f) or None)
            ]
            if diff_fields:
                conflicts.append({
                    "row": excel_row_no,
                    "existing_id": existing.get("id"),
                    "diff_fields": diff_fields,
                    "existing": {f: existing.get(f) for f in diff_fields},
                    "incoming": {f: doc.get(f) for f in diff_fields},
                })
        elif existing_by_key is not None:
            news.append({
                "row": excel_row_no,
                "make": doc.get("make"),
                "model": doc.get("model"),
                "variant": doc.get("variant"),
                "year_generation": doc.get("year_generation"),
                "front_tyre_size": doc.get("front_tyre_size"),
                "rear_tyre_size": doc.get("rear_tyre_size"),
            })

        prepared.append({"natural_key": list(key), "doc": doc})

    return {
        "ok": True,
        "sheet": sheet_name,
        "excel_rows_on_disk": max(0, len(rows) - 1),
        "prepared_rows": prepared,
        "invalid_rows": invalid,
        "conflicts": conflicts,
        "duplicates": duplicates,
        "new_rows": news,
        "counts": {
            "excel_rows_on_disk": max(0, len(rows) - 1),
            "prepared": len(prepared),
            "invalid": len(invalid),
            "conflicts": len(conflicts),
            "duplicates_within_file": len(duplicates),
            "new_rows": len(news) if existing_by_key is not None else None,
        },
    }
